"""
Großer Benchmark Parser
========================
Actual structure (confirmed from inspection):

File: TemplateComparisonAnalytics.xlsx
- 99 sheets total
- Data sheets named: "{PROJECT} {TEMPLATE}"
  Projects : FLEX, ECSS_E60-30, CS_E 50, TSS, EVS  (+ Random1..5)
  Templates: free, EARS, MASTER, advEARS, boilerplates (DODT), SPIDER

- Header row: row 15 (0-indexed row 14)
  Columns: ID, Text, template, liability?(R5), structured_sentence?(R6), ...

- Data starts: row 16 (0-indexed row 15)

- The "template" column (col 3) in EARS/MASTER sheets contains:
  EARS type labels like "Ubiquitous requirement", "Event-driven", etc.

Strategy for our paper:
  - Extract (free_text, EARS_text, MASTER_text) aligned triples per requirement ID
  - Focus on 5 real projects (exclude Random groups)
  - For us: free → target norm pairs are the base items

Run:
    python parse_grosser.py \
        --input  /path/to/TemplateComparisonAnalytics.xlsx \
        --output-dir /path/to/parsed/

Outputs:
    grosser_requirements.json   — aligned triples
    grosser_report.json         — stats
"""

import json
import argparse
import zipfile
import io
from pathlib import Path
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    raise SystemExit("Run: pip install openpyxl")


# ── Projects and templates we care about ──────────────────────────────────
REAL_PROJECTS = {"FLEX", "ECSS_E60-30", "CS_E 50", "TSS", "EVS"}
ALL_TEMPLATES  = {"free", "EARS", "MASTER", "advEARS", "boilerplates (DODT)", "SPIDER"}
CORE_TEMPLATES = {"free", "EARS", "MASTER"}   # what we extract for the paper

DATA_START_ROW = 15   # 0-indexed (row 16 in Excel)
HEADER_ROW     = 14   # 0-indexed (row 15 in Excel)


def parse_sheet_name(name: str) -> tuple[str, str] | None:
    """
    Split 'FLEX EARS' → ('FLEX', 'EARS')
    Handles multi-word templates like 'boilerplates (DODT)'
    Returns None for non-data sheets (Summary, Overview, etc.)
    """
    for template in sorted(ALL_TEMPLATES, key=len, reverse=True):
        if name.endswith(template):
            project = name[: -len(template)].strip()
            if project:
                return project, template
    return None


def extract_sheet_data(ws) -> list[dict]:
    """
    Extract requirement rows from one sheet.
    Returns list of dicts with: id, text, template_label, row_num
    """
    rows = list(ws.iter_rows(values_only=True))
    records = []

    for i, row in enumerate(rows):
        if i < DATA_START_ROW:
            continue

        req_id   = row[0] if len(row) > 0 else None
        text     = row[1] if len(row) > 1 else None
        tmpl_lbl = row[2] if len(row) > 2 else None  # EARS type label

        # Skip empty rows
        if req_id is None and text is None:
            continue

        # Skip formula-only rows (text is a formula string)
        if isinstance(text, str) and text.startswith("="):
            continue

        # Skip if text is None or too short
        if not text or (isinstance(text, str) and len(text.split()) < 3):
            continue

        records.append({
            "req_id":        req_id,
            "text":          str(text).strip(),
            "template_label": str(tmpl_lbl).strip() if tmpl_lbl else None,
            "word_count":    len(str(text).split()),
        })

    return records


def build_aligned_pairs(workbook, projects: set) -> list[dict]:
    """
    Build aligned (free_text, EARS_text, MASTER_text) triples
    per project and requirement ID.
    """
    # Collect all sheet data: {project: {template: {req_id: record}}}
    sheet_data = defaultdict(lambda: defaultdict(dict))

    sheet_stats = []
    for sheet_name in workbook.sheetnames:
        parsed = parse_sheet_name(sheet_name)
        if not parsed:
            continue

        project, template = parsed
        if project not in projects:
            continue
        if template not in CORE_TEMPLATES:
            continue

        ws = workbook[sheet_name]
        records = extract_sheet_data(ws)
        sheet_stats.append((sheet_name, len(records)))

        for r in records:
            req_id = r["req_id"]
            if req_id is not None:
                sheet_data[project][template][req_id] = r

        print(f"  [{sheet_name:<25}] {len(records):>3} requirements")

    # Build aligned triples
    aligned = []
    for project in sorted(sheet_data.keys()):
        free_reqs = sheet_data[project].get("free", {})
        ears_reqs = sheet_data[project].get("EARS", {})
        mstr_reqs = sheet_data[project].get("MASTER", {})

        # Use free requirements as anchor
        for req_id, free_rec in free_reqs.items():
            ears_rec = ears_reqs.get(req_id)
            mstr_rec = mstr_reqs.get(req_id)

            record = {
                "project":              project,
                "requirement_id":       f"{project}_{req_id}",
                "req_id_raw":           req_id,
                # Free text (source for revision)
                "free_text":            free_rec["text"],
                "free_word_count":      free_rec["word_count"],
                # EARS variant
                "ears_text":            ears_rec["text"] if ears_rec else None,
                "ears_template_label":  ears_rec["template_label"] if ears_rec else None,
                "ears_word_count":      ears_rec["word_count"] if ears_rec else None,
                # MASTER variant
                "master_text":          mstr_rec["text"] if mstr_rec else None,
                "master_template_label":mstr_rec["template_label"] if mstr_rec else None,
                "master_word_count":    mstr_rec["word_count"] if mstr_rec else None,
                # Alignment status
                "has_ears":             ears_rec is not None,
                "has_master":           mstr_rec is not None,
                "fully_aligned":        ears_rec is not None and mstr_rec is not None,
            }
            aligned.append(record)

    return aligned


def print_report(aligned: list[dict]):
    project_dist = Counter(r["project"] for r in aligned)
    fully_aligned = sum(1 for r in aligned if r["fully_aligned"])
    ears_only = sum(1 for r in aligned if r["has_ears"] and not r["has_master"])
    master_only = sum(1 for r in aligned if r["has_master"] and not r["has_ears"])

    # EARS template label distribution
    ears_labels = Counter(
        r["ears_template_label"]
        for r in aligned
        if r["ears_template_label"]
    )

    print(f"\n{'='*60}")
    print(f"GROẞER BENCHMARK — Parsed Results")
    print(f"{'='*60}")
    print(f"  Total free requirements : {len(aligned)}")
    print(f"  Fully aligned (F+E+M)   : {fully_aligned}")
    print(f"  EARS only               : {ears_only}")
    print(f"  MASTER only             : {master_only}")

    print(f"\n  Per project:")
    for proj, cnt in sorted(project_dist.items()):
        proj_recs = [r for r in aligned if r["project"] == proj]
        fa = sum(1 for r in proj_recs if r["fully_aligned"])
        print(f"    {proj:<20} {cnt:>3} reqs  ({fa} fully aligned)")

    print(f"\n  EARS template label distribution:")
    for label, cnt in sorted(ears_labels.items(), key=lambda x: -x[1])[:10]:
        print(f"    {label:<35} {cnt}")

    print(f"\n  Word count comparison (free vs EARS vs MASTER):")
    free_wc  = [r["free_word_count"] for r in aligned]
    ears_wc  = [r["ears_word_count"] for r in aligned if r["ears_word_count"]]
    mstr_wc  = [r["master_word_count"] for r in aligned if r["master_word_count"]]

    def stats(wc):
        if not wc: return "N/A"
        return f"mean={sum(wc)/len(wc):.1f} median={sorted(wc)[len(wc)//2]}"

    print(f"    Free   : {stats(free_wc)}")
    print(f"    EARS   : {stats(ears_wc)}")
    print(f"    MASTER : {stats(mstr_wc)}")

    print(f"\n  Sample aligned triple:")
    sample = next((r for r in aligned if r["fully_aligned"]), None)
    if sample:
        print(f"    [FREE  ] {sample['free_text'][:100]}")
        print(f"    [EARS  ] {sample['ears_text'][:100]}")
        print(f"    [MASTER] {sample['master_text'][:100]}")


def run(input_path: Path, output_dir: Path):
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"GROẞER BENCHMARK PARSER")
    print(f"Input  : {input_path}")
    print(f"{'='*60}")

    # Handle both direct xlsx and zip input
    if input_path.suffix == ".zip":
        print("  Extracting xlsx from zip...")
        with zipfile.ZipFile(input_path) as zf:
            xlsx_names = [n for n in zf.namelist() if n.endswith("TemplateComparisonAnalytics.xlsx")]
            if not xlsx_names:
                raise FileNotFoundError("TemplateComparisonAnalytics.xlsx not found in zip")
            with zf.open(xlsx_names[0]) as f:
                wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"\n  Extracting sheets...")

    aligned = build_aligned_pairs(wb, REAL_PROJECTS)
    print_report(aligned)

    # Save
    out_json = output_dir / "grosser_requirements.json"
    out_json.write_text(json.dumps(aligned, indent=2, ensure_ascii=False))
    print(f"\n  Saved → {out_json}  ({len(aligned)} records)")

    report = {
        "dataset":              "GrosserBenchmark",
        "total_free_reqs":      len(aligned),
        "fully_aligned":        sum(1 for r in aligned if r["fully_aligned"]),
        "has_ears":             sum(1 for r in aligned if r["has_ears"]),
        "has_master":           sum(1 for r in aligned if r["has_master"]),
        "project_distribution": dict(Counter(r["project"] for r in aligned)),
        "ears_label_distribution": dict(Counter(
            r["ears_template_label"] for r in aligned if r["ears_template_label"]
        )),
    }
    out_report = output_dir / "grosser_report.json"
    out_report.write_text(json.dumps(report, indent=2, ensure_ascii=False))
    print(f"  Saved → {out_report}")


def main():
    parser = argparse.ArgumentParser(description="Parse Großer benchmark")
    parser.add_argument("--input",      required=True,
                        help="Path to TemplateComparisonAnalytics.xlsx OR the .zip file")
    parser.add_argument("--output-dir", required=True)
    args = parser.parse_args()
    run(Path(args.input), Path(args.output_dir))


if __name__ == "__main__":
    main()
